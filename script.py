from google.cloud import bigquery
from sqlalchemy import and_, or_
from common.constants import SubscriptionStatus
import openpyxl
from common.clients.license_generator.constants import ProtectionCodeType
from common.subscription import SubscriptionOperations

def get_cct_subscriptions() -> None:

    START_DT = '2020-08-01'
    END_DT = '2020-09-30'
    CCT_MODEL_ID = 21
    LIMIT = 60000
    OFFSET = 150000

    client = bigquery.Client()

    # Находим все ККТ
    ccts = mdl.Hub.query.outerjoin(
        mdl.CctModel,
        mdl.Hub.cct_model_id == mdl.CctModel.id
    ).outerjoin(
        mdl.CctMeta,
        mdl.CctMeta.hub_id == mdl.Hub.id,
    ).filter(
        mdl.CctModel.id == CCT_MODEL_ID,
    ).order_by(mdl.Hub.id).limit(LIMIT).offset(OFFSET).all()
    ccts_dict = {
        x.serial_id: x for x in ccts
    }

    # Смотрим, какие из них оплачены
    service = mdl.Service.query.filter(
        and_(
            mdl.Service.code == mdl.Service.CODE_CCT_FIRMWARE_OFFLINE,
            mdl.Service.platform == common_constants.Platform.P_50,
        )
    ).first()
    paid_info = SubscriptionOperations.get_service_paid_info_bulk(
        service=service, serial_ids=set(ccts_dict.keys()),
    )

    # Собираем все коды защиты
    protection_codes = mdl.ProtectionCode.query.with_entities(
        mdl.ProtectionCode.serial_id
    ).filter(
        and_(
            mdl.ProtectionCode.delivery_status.isnot(None),
            mdl.ProtectionCode.type == ProtectionCodeType.P_10,
            mdl.ProtectionCode.serial_id.in_(set(ccts_dict.keys())),
        )
    )
    protection_codes_set = {x.serial_id for x in protection_codes}

    # строка с серийниками для фильтра в запросе в бд
    ccts_query = "','".join([x for x in ccts_dict])
    print('отправляем запрос')
    query = f"""
        SELECT serial_id, max(resc) - min(resc) as resc, sum(case when efn > 0 then 1 else 0 end) as error_count
        FROM `atol-accountapi.cct_statistics.prod`
        WHERE DATE(backend_timestamp) BETWEEN "{START_DT}" AND "{END_DT}"
        and serial_id IN ('{ccts_query}')
        GROUP BY serial_id
    """
    query_job = client.query(query)

    # Куда сохранять файл
    filepath = f'/Users/vadimvadim/Documents/{OFFSET}-{OFFSET+LIMIT}.xlsx'
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.cell(row=1, column=1, value='Период')
    sheet.cell(row=1, column=2, value='*РН ККТ')
    sheet.cell(row=1, column=3, value='Код модели ККТ')
    sheet.cell(row=1, column=4, value='*Наименование модели ККТ')
    sheet.cell(row=1, column=5, value='*ЗН ККТ')
    sheet.cell(row=1, column=6, value='*Корректность лицензии (0-нет/1-есть)')
    sheet.cell(row=1, column=7, value='Кол-во отрезов за период')
    sheet.cell(row=1, column=8, value='Кол-во ошибок ФН за период')
    # sheet.cell(row=1, column=9, value='Дата окончания лицензии')

    for index, row in enumerate(query_job):
        is_license = 0
        if row['serial_id'] in protection_codes_set:
            is_license = 1
        elif paid_info[row['serial_id']]['is_paid']:
            is_license = 1
        sheet.cell(row=index+2, column=1, value=f'{START_DT}-{END_DT}')
        sheet.cell(row=index+2, column=2, value=ccts_dict[row['serial_id']].cct_meta.rnm)
        sheet.cell(row=index+2, column=3, value=ccts_dict[row['serial_id']].cct_model.code)
        sheet.cell(row=index+2, column=4, value=ccts_dict[row['serial_id']].cct_model.name)
        sheet.cell(row=index+2, column=5, value=row['serial_id'])
        sheet.cell(row=index+2, column=6, value=is_license)
        sheet.cell(row=index+2, column=7, value=row['resc'])
        sheet.cell(row=index+2, column=8, value=row['error_count'])
    wb.save(filepath)

    return None
